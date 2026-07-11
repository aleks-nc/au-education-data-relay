#!/usr/bin/env python3
"""
AU education data relay — runs on GitHub Actions (unrestricted network).

Downloads the government spreadsheets the dashboard sandbox cannot reach,
parses what it can into data/series_bundle.json (dashboard-schema series),
dumps EVERY sheet of every workbook to data/csv/<source>/<sheet>.csv so the
data is always accessible even when structured parsing fails, and writes a
per-source status manifest. Never raises: one failing source must not block
the others.
"""
import datetime as dt
import io
import json
import re
import traceback
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
CSV = DATA / "csv"
RAW = ROOT / "raw"
for p in (DATA, CSV, RAW):
    p.mkdir(parents=True, exist_ok=True)

UA = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36"}
RAW_SIZE_CAP = 20 * 1024 * 1024  # commit raw files up to 20 MB

manifest = {"generated_at": dt.datetime.now(dt.timezone.utc).isoformat(), "sources": {}}
bundle = {"generated_at": manifest["generated_at"], "datasets": []}


def log(src, **kw):
    manifest["sources"].setdefault(src, {}).update(kw)
    print(f"[{src}] " + json.dumps(kw, default=str)[:300])


def get(url, binary=False, timeout=120, tries=3):
    last = None
    for attempt in range(tries):
        try:
            r = requests.get(url, headers=UA, timeout=timeout, allow_redirects=True)
            r.raise_for_status()
            return r.content if binary else r.text
        except Exception as e:
            last = e
            import time
            time.sleep(10 * (attempt + 1))
    raise last


def save_raw(src, filename, content):
    if len(content) <= RAW_SIZE_CAP:
        d = RAW / src
        d.mkdir(parents=True, exist_ok=True)
        (d / filename).write_bytes(content)


def dump_workbook(src, filename, content):
    """Dump every sheet to CSV; return sheet inventory."""
    inv = []
    try:
        xls = pd.ExcelFile(io.BytesIO(content))
    except Exception as e:
        return [{"error": f"unreadable workbook: {e}"}]
    outdir = CSV / src
    outdir.mkdir(parents=True, exist_ok=True)
    stem = re.sub(r"[^A-Za-z0-9_-]+", "_", Path(filename).stem)[:60]
    for sheet in xls.sheet_names:
        try:
            df = xls.parse(sheet, header=None)
            safe = re.sub(r"[^A-Za-z0-9_-]+", "_", sheet)[:40]
            df.to_csv(outdir / f"{stem}__{safe}.csv", index=False, header=False)
            inv.append({"sheet": sheet, "rows": len(df), "cols": df.shape[1],
                        "head": df.head(3).fillna("").astype(str).values.tolist()})
        except Exception as e:
            inv.append({"sheet": sheet, "error": str(e)})
    return inv


def xlsx_links(page_url, host=None):
    html = get(page_url)
    soup = BeautifulSoup(html, "lxml")
    out = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if re.search(r"\.(xlsx?|csv)(\?|$)", href, re.I):
            if href.startswith("/"):
                m = re.match(r"(https?://[^/]+)", host or page_url)
                href = m.group(1) + href
            out.append({"url": href, "text": a.get_text(" ", strip=True)[:120]})
    return out


# ------------------------- 1. ABS Labour Force (known structure) -----------
def fetch_abs():
    src = "abs_labour_force"
    try:
        page = "https://www.abs.gov.au/statistics/labour/employment-and-unemployment/labour-force-australia/latest-release"
        links = xlsx_links(page, host="https://www.abs.gov.au")
        t1 = next((l for l in links if re.search(r"6202001|Table 1", l["text"] + l["url"], re.I)), None)
        log(src, page=page, links_found=len(links), table1=bool(t1))
        if not t1:
            return
        content = get(t1["url"], binary=True)
        save_raw(src, "6202_table1.xlsx", content)
        inv = dump_workbook(src, "6202_table1", content)
        log(src, sheets=[i.get("sheet") for i in inv])
        # Data1 sheet: row with 'Series ID' in col A marks the header; A84423050A = unemployment rate SA
        xls = pd.ExcelFile(io.BytesIO(content))
        for sheet in ("Data1", "Data2"):
            if sheet not in xls.sheet_names:
                continue
            df = xls.parse(sheet, header=None)
            hdr = df.index[df[0].astype(str).str.strip() == "Series ID"]
            if len(hdr) == 0:
                continue
            h = hdr[0]
            cols = df.iloc[h].astype(str).tolist()
            if "A84423050A" not in cols:
                continue
            c = cols.index("A84423050A")
            sub = df.iloc[h + 1:, [0, c]].dropna()
            sub.columns = ["date", "rate"]
            sub["date"] = pd.to_datetime(sub["date"], errors="coerce")
            sub = sub.dropna().tail(48)  # last 4 years
            periods = [d.strftime("%b %Y") for d in sub["date"]]
            values = [round(float(v), 1) for v in sub["rate"]]
            bundle["datasets"].append({
                "id": "abs_unemployment_monthly", "title": "Unemployment rate — monthly, seasonally adjusted (%)",
                "frequency": "monthly", "unit": "%", "agg": "mean",
                "source": "ABS Labour Force Australia, table 1 (series A84423050A)",
                "source_url": page, "confidence": "high (official ABS spreadsheet, current vintage)",
                "periods": periods, "series": [{"name": "Unemployment rate %", "values": values}]})
            log(src, parsed="abs_unemployment_monthly", points=len(values), latest=f"{periods[-1]}={values[-1]}")
            return
        log(src, parse_warning="series A84423050A not located; sheets dumped to CSV")
    except Exception as e:
        log(src, error=str(e), trace=traceback.format_exc()[-400:])


# ------------- 2. DoE international student data (direct file URLs) --------
DOE_FILES = [
    ("may-2026-all-data", "https://www.education.gov.au/download/20217/international-student-data-year-date-ytd/45085/may-2026-all-data/xlsx"),
    ("may-2026-latest-data", "https://www.education.gov.au/download/20217/international-student-data-year-date-ytd/45084/may-2026-latest-data/xlsx"),
    ("december-2025-all-data", "https://www.education.gov.au/download/20217/international-student-data-year-date-ytd/44305/december-2025-all-data/xlsx"),
]


def inventory_big_workbook(src, name, content, cap_rows=300):
    """Memory-safe inventory + capped CSV dump using openpyxl read_only streaming."""
    import csv
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    outdir = CSV / src
    outdir.mkdir(parents=True, exist_ok=True)
    inv = []
    for ws in wb.worksheets:
        rows_dumped = 0
        head = []
        path = outdir / f"{name}__{re.sub(r'[^A-Za-z0-9_-]+', '_', ws.title)[:40]}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            total = 0
            for row in ws.iter_rows(values_only=True):
                total += 1
                if total <= cap_rows:
                    vals = ["" if v is None else str(v) for v in row]
                    w.writerow(vals)
                    rows_dumped += 1
                    if total <= 6:
                        head.append(vals[:12])
            inv.append({"sheet": ws.title, "total_rows": total, "dumped_rows": rows_dumped, "head": head})
    wb.close()
    return inv


def fetch_doe_monthly():
    src = "doe_monthly_summary"
    for name, url in DOE_FILES:
        try:
            content = get(url, binary=True, timeout=600, tries=3)
            log(src, **{f"{name}_bytes": len(content)})
            save_raw(src, f"{name}.xlsx", content)  # skipped automatically if > cap
            inv = inventory_big_workbook(src, name, content)
            log(src, **{f"{name}_inventory": inv})
        except Exception as e:
            log(src, **{f"{name}_error": str(e)})


# ------------- 3. data.gov.au — student visas (grants by month etc.) -------
def fetch_datagov(package_query, src, keep=6):
    try:
        r = None
        for api in ("https://data.gov.au/data/api/3/action/package_search",
                    "https://data.gov.au/api/3/action/package_search"):
            try:
                resp = requests.get(api, params={"q": package_query, "rows": 5},
                                    headers={**UA, "Accept": "application/json"}, timeout=60)
                if "json" in resp.headers.get("content-type", ""):
                    r = resp.json()
                    break
                log(src, **{f"ckan_nonjson_{api[-30:]}": resp.text[:150]})
            except Exception as e:
                log(src, **{f"ckan_err_{api[-30:]}": str(e)})
        if r is None:
            return
        results = r.get("result", {}).get("results", [])
        log(src, query=package_query, packages=[p["name"] for p in results])
        if not results:
            return
        pkg = results[0]
        resources = pkg.get("resources", [])
        picked = [res for res in resources if re.search(r"\.(xlsx?|csv)$", res.get("url", ""), re.I)]
        picked.sort(key=lambda r: r.get("last_modified") or r.get("created") or "", reverse=True)
        log(src, resources=[{"name": res.get("name", "")[:80], "modified": res.get("last_modified")} for res in picked[:keep]])
        for res in picked[:keep]:
            try:
                content = get(res["url"], binary=True)
                fname = re.sub(r"[^A-Za-z0-9._-]+", "_", res["url"].split("/")[-1])
                save_raw(src, fname, content)
                if fname.lower().endswith((".xlsx", ".xls")):
                    dump_workbook(src, fname, content)
                else:
                    (CSV / src).mkdir(parents=True, exist_ok=True)
                    (CSV / src / fname).write_bytes(content)
            except Exception as e:
                log(src, **{f"resource_error_{res.get('name','')[:40]}": str(e)})
    except Exception as e:
        log(src, error=str(e), trace=traceback.format_exc()[-400:])


# ----------------------------- 4. NCVER ------------------------------------
def fetch_ncver():
    src = "ncver"
    try:
        for page in (
            "https://www.ncver.edu.au/research-and-statistics/publications/all-publications/apprentices-and-trainees-2025-december-quarter",
            "https://www.ncver.edu.au/research-and-statistics/collections/apprentices-and-trainees-collection",
        ):
            try:
                links = xlsx_links(page, host="https://www.ncver.edu.au")
                log(src, **{f"links_{page.rsplit('/', 1)[-1][:40]}": links[:8]})
                for l in links[:3]:
                    content = get(l["url"], binary=True)
                    fname = re.sub(r"[^A-Za-z0-9._-]+", "_", l["url"].split("/")[-1].split("?")[0])
                    save_raw(src, fname, content)
                    dump_workbook(src, fname, content)
            except Exception as e:
                log(src, **{f"page_error_{page[-30:]}": str(e)})
    except Exception as e:
        log(src, error=str(e))


def main():
    fetch_abs()
    fetch_doe_monthly()
    fetch_datagov("student visas", "homeaffairs_student_visas")
    fetch_datagov("temporary entrants visa holders", "homeaffairs_temp_entrants")
    fetch_ncver()
    (DATA / "manifest.json").write_text(json.dumps(manifest, indent=1, default=str), encoding="utf-8")
    (DATA / "series_bundle.json").write_text(json.dumps(bundle, indent=1), encoding="utf-8")
    print(f"\nDone. Bundle datasets: {len(bundle['datasets'])}. "
          f"Sources: {list(manifest['sources'])}")


if __name__ == "__main__":
    main()
